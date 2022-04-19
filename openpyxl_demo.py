import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
print(cell.row)
print(cell.column)
print(cell.coordinate)

cell = sheet["a1"]
cell = sheet.cell(row=1, column=1)

print(sheet.max_row)
print(sheet.max_column)



sheet.append([103, "2/1/2020", 99]) # used to add a row, info needs to be given as tuple

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column +1):
        cell = sheet.cell(row, col)
        print(cell.value)

wb.save("transaction2.xlsx") # will save in new updated file